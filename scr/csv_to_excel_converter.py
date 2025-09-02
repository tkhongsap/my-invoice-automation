#!/usr/bin/env python3
"""
CSV to Excel Converter

This script converts the consolidated invoice CSV file to a formatted Excel file.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def setup_paths():
    """Set up input and output file paths."""
    base_dir = Path(__file__).parent.parent
    csv_path = base_dir / "output" / "consolidated_invoices.csv"
    excel_path = base_dir / "output" / "consolidated_invoices.xlsx"
    
    if not csv_path.exists():
        print(f"Error: CSV file not found at {csv_path}")
        return None, None
    
    return csv_path, excel_path

def format_excel_file(excel_path):
    """Apply formatting to the Excel file."""
    # Load the workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Apply borders and alignment to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border_style
            # Right-align amount column
            if cell.column == 3:  # Amount column
                cell.alignment = Alignment(horizontal="right")
                # Format as number with 2 decimal places
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Set minimum and maximum widths
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Specific column width adjustments
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 45  # Description
    ws.column_dimensions['C'].width = 15  # Amount
    ws.column_dimensions['D'].width = 40  # Source File
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add a summary section
    summary_start_row = ws.max_row + 3
    
    # Add summary headers
    ws.cell(row=summary_start_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=summary_start_row + 1, column=1, value="Total Transactions:").font = Font(bold=True)
    ws.cell(row=summary_start_row + 1, column=2, value=ws.max_row - 1)
    
    ws.cell(row=summary_start_row + 2, column=1, value="Total Amount (THB):").font = Font(bold=True)
    total_formula = f"=SUM(C2:C{ws.max_row - 2})"
    ws.cell(row=summary_start_row + 2, column=2, value=total_formula)
    ws.cell(row=summary_start_row + 2, column=2).number_format = '#,##0.00'
    ws.cell(row=summary_start_row + 2, column=2).font = Font(bold=True)
    
    # Save the formatted workbook
    wb.save(excel_path)
    return True

def add_vendor_summary_sheet(excel_path, df):
    """Add a vendor summary sheet to the Excel file."""
    # Group by vendor (extract vendor name from description)
    df['Vendor'] = df['Description'].str.extract(r'^"?([^,\d]+)')
    df['Vendor'] = df['Vendor'].str.strip()
    
    # Create vendor summary
    vendor_summary = df.groupby('Vendor').agg({
        'Amount (THB)': ['sum', 'count', 'mean']
    }).round(2)
    
    vendor_summary.columns = ['Total Amount (THB)', 'Transaction Count', 'Average Amount (THB)']
    vendor_summary = vendor_summary.sort_values('Total Amount (THB)', ascending=False)
    vendor_summary.reset_index(inplace=True)
    
    # Load workbook and add vendor summary sheet
    wb = load_workbook(excel_path)
    
    # Create vendor summary sheet
    if 'Vendor Summary' in wb.sheetnames:
        del wb['Vendor Summary']
    
    ws_summary = wb.create_sheet('Vendor Summary')
    
    # Write headers
    headers = ['Vendor', 'Total Amount (THB)', 'Transaction Count', 'Average Amount (THB)']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for idx, row in vendor_summary.iterrows():
        ws_summary.cell(row=idx+2, column=1, value=row['Vendor'])
        ws_summary.cell(row=idx+2, column=2, value=row['Total Amount (THB)']).number_format = '#,##0.00'
        ws_summary.cell(row=idx+2, column=3, value=row['Transaction Count'])
        ws_summary.cell(row=idx+2, column=4, value=row['Average Amount (THB)']).number_format = '#,##0.00'
    
    # Apply borders
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws_summary.iter_rows(min_row=1, max_row=len(vendor_summary)+1):
        for cell in row:
            cell.border = border_style
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 20
    
    # Save workbook
    wb.save(excel_path)

def main():
    """Main function to convert CSV to Excel."""
    print("CSV to Excel Converter")
    print("=" * 60)
    
    # Setup paths
    csv_path, excel_path = setup_paths()
    if not csv_path:
        return
    
    print(f"Reading CSV from: {csv_path}")
    
    # Read CSV file
    df = pd.read_csv(csv_path)
    
    # Convert Date column to datetime
    df['Date'] = pd.to_datetime(df['Date'])
    
    # Save to Excel
    print(f"Converting to Excel: {excel_path}")
    df.to_excel(excel_path, index=False, sheet_name='Transactions')
    
    # Apply formatting
    print("Applying formatting...")
    if format_excel_file(excel_path):
        print("Formatting applied successfully")
    
    # Add vendor summary sheet
    print("Adding vendor summary sheet...")
    add_vendor_summary_sheet(excel_path, df)
    
    print("\n" + "=" * 60)
    print(f"Excel file created successfully!")
    print(f"Output saved to: {excel_path}")
    print(f"Total transactions: {len(df)}")
    print(f"Total amount: ฿{df['Amount (THB)'].sum():,.2f}")
    print("=" * 60)

if __name__ == "__main__":
    main()