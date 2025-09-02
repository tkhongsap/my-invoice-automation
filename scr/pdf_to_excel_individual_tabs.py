#!/usr/bin/env python3
"""
PDF to Excel Individual Tabs

This script creates individual tabs in Excel for each invoice,
with a full-page screenshot clearly visible in each tab.
"""

import os
import sys
from pathlib import Path
import fitz  # PyMuPDF
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re

def setup_paths():
    """Set up input and output paths."""
    base_dir = Path(__file__).parent.parent
    invoices_dir = base_dir / "invoices"
    excel_path = base_dir / "output" / "consolidated_invoices.xlsx"
    csv_path = base_dir / "output" / "consolidated_invoices.csv"
    
    if not invoices_dir.exists():
        print(f"Error: Invoices directory not found at {invoices_dir}")
        return None, None, None
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        return None, None, None
    
    return invoices_dir, excel_path, csv_path

def extract_pdf_full_page(pdf_path, dpi=200):
    """Extract a full-page screenshot from PDF at high quality."""
    try:
        # Open PDF
        doc = fitz.open(pdf_path)
        
        # Get first page
        page = doc[0]
        
        # Render page to image at specified DPI for clarity
        mat = fitz.Matrix(dpi/72.0, dpi/72.0)
        pix = page.get_pixmap(matrix=mat)
        
        # Convert to PIL Image
        img_data = pix.pil_tobytes(format="PNG")
        img = Image.open(io.BytesIO(img_data))
        
        # Resize to fit Excel viewport nicely (max width 800px for clear viewing)
        max_width = 800
        if img.width > max_width:
            ratio = max_width / img.width
            new_height = int(img.height * ratio)
            img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
        
        doc.close()
        
        return img
        
    except Exception as e:
        print(f"Error extracting screenshot from {pdf_path.name}: {str(e)}")
        return None

def create_invoice_tab(wb, pdf_file, invoices_dir, tab_number, transaction_data):
    """Create a single tab for an invoice with its screenshot."""
    
    # Create sheet name (e.g., "Inv_01" for better tab management)
    sheet_name = f"Inv_{tab_number:02d}"
    
    # Remove sheet if it already exists
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    # Create new sheet
    ws = wb.create_sheet(sheet_name)
    
    # Add header information
    header_font = Font(bold=True, size=14)
    info_font = Font(bold=True, size=11)
    
    # Title
    ws.cell(row=1, column=1, value=f"Invoice #{tab_number:02d}").font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Transaction details
    ws.cell(row=3, column=1, value="Date:").font = info_font
    ws.cell(row=3, column=2, value=transaction_data['Date'])
    
    ws.cell(row=4, column=1, value="Description:").font = info_font
    ws.cell(row=4, column=2, value=transaction_data['Description'])
    ws.merge_cells('B4:D4')
    
    ws.cell(row=5, column=1, value="Amount (THB):").font = info_font
    amount_cell = ws.cell(row=5, column=2, value=transaction_data['Amount (THB)'])
    amount_cell.number_format = '#,##0.00'
    
    ws.cell(row=6, column=1, value="Source File:").font = info_font
    ws.cell(row=6, column=2, value=transaction_data['Source File'])
    ws.merge_cells('B6:D6')
    
    # Add separator line
    ws.cell(row=8, column=1, value="Invoice Image:").font = Font(bold=True, size=12)
    
    # Set column widths for better layout
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Extract and add screenshot
    pdf_path = invoices_dir / pdf_file
    
    if pdf_path.exists():
        # Extract full-page screenshot
        img = extract_pdf_full_page(pdf_path)
        
        if img:
            # Save to temporary BytesIO object
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Create openpyxl Image object
            xl_img = XLImage(img_buffer)
            
            # Position image starting at row 10 to leave space for header info
            ws.add_image(xl_img, 'A10')
            
            # Add image dimensions info
            ws.cell(row=8, column=2, value=f"({img.width}x{img.height}px)")
            
            return True, img.width, img.height
        else:
            ws.cell(row=10, column=1, value="Error loading image")
            return False, 0, 0
    else:
        ws.cell(row=10, column=1, value="PDF file not found")
        return False, 0, 0

def add_index_sheet(wb, df):
    """Add or update an index sheet with links to all invoice tabs."""
    
    # Remove existing index sheet if present
    if 'Invoice Index' in wb.sheetnames:
        del wb['Invoice Index']
    
    # Create index sheet at the beginning
    ws = wb.create_sheet('Invoice Index', 0)
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    title_cell = ws.cell(row=1, column=1, value="Invoice Index - Quick Navigation")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    title_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['Invoice #', 'Tab Name', 'Date', 'Description', 'Amount (THB)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Add data with hyperlinks to tabs
    for idx, row in df.iterrows():
        row_num = idx + 4
        tab_number = idx + 1
        tab_name = f"Inv_{tab_number:02d}"
        
        # Invoice number
        ws.cell(row=row_num, column=1, value=tab_number).border = border_style
        
        # Tab name with hyperlink
        link_cell = ws.cell(row=row_num, column=2, value=tab_name)
        link_cell.hyperlink = f"#'{tab_name}'!A1"
        link_cell.font = Font(color="0000FF", underline="single")
        link_cell.border = border_style
        
        # Date
        ws.cell(row=row_num, column=3, value=row['Date']).border = border_style
        
        # Description
        ws.cell(row=row_num, column=4, value=row['Description']).border = border_style
        
        # Amount
        amount_cell = ws.cell(row=row_num, column=5, value=row['Amount (THB)'])
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border_style
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Add summary
    summary_row = len(df) + 5
    ws.cell(row=summary_row, column=1, value="Total:").font = Font(bold=True)
    total_cell = ws.cell(row=summary_row, column=5, value=df['Amount (THB)'].sum())
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True)

def main():
    """Main function to add individual invoice tabs to Excel."""
    print("PDF to Excel - Individual Invoice Tabs")
    print("=" * 60)
    
    # Setup paths
    invoices_dir, excel_path, csv_path = setup_paths()
    if not invoices_dir:
        sys.exit(1)
    
    # Read transaction data
    df = pd.read_csv(csv_path)
    
    # Load existing Excel file
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path)
    
    # Remove old "Invoice Screenshots" sheet if it exists
    if 'Invoice Screenshots' in wb.sheetnames:
        del wb['Invoice Screenshots']
        print("Removed old 'Invoice Screenshots' sheet")
    
    print(f"\nCreating {len(df)} individual invoice tabs...")
    print("=" * 60)
    
    successful = 0
    failed = 0
    
    # Process each invoice
    for idx, row in df.iterrows():
        tab_number = idx + 1
        pdf_file = row['Source File']
        
        print(f"[{tab_number}/{len(df)}] Creating tab for {pdf_file}...")
        
        # Create invoice tab
        success, width, height = create_invoice_tab(
            wb, pdf_file, invoices_dir, tab_number, row
        )
        
        if success:
            successful += 1
            print(f"  ✓ Tab 'Inv_{tab_number:02d}' created ({width}x{height}px)")
        else:
            failed += 1
            print(f"  ✗ Failed to create tab")
    
    # Add index sheet
    print("\nCreating Invoice Index sheet...")
    add_index_sheet(wb, df)
    print("✓ Index sheet created with navigation links")
    
    # Save the workbook
    print("\nSaving Excel file...")
    wb.save(excel_path)
    
    print("\n" + "=" * 60)
    print("✓ Excel file updated successfully!")
    print(f"✓ File: {excel_path}")
    print(f"✓ Created {successful} invoice tabs")
    if failed > 0:
        print(f"⚠ Failed to create {failed} tabs")
    
    print("\nFeatures added:")
    print("  • Invoice Index sheet with navigation links")
    print(f"  • {successful} individual invoice tabs (Inv_01 to Inv_{len(df):02d})")
    print("  • Each tab contains full invoice screenshot")
    print("  • Transaction details at the top of each tab")
    print("  • High-quality images (800px width) for clear viewing")
    print("=" * 60)

if __name__ == "__main__":
    main()