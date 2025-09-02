#!/usr/bin/env python3
"""
PDF to Excel Screenshots

This script extracts screenshots from PDF invoices and adds them to the Excel file
as a reference sheet with embedded images.
"""

import os
import sys
from pathlib import Path
import fitz  # PyMuPDF
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def setup_paths():
    """Set up input and output paths."""
    base_dir = Path(__file__).parent.parent
    invoices_dir = base_dir / "invoices"
    excel_path = base_dir / "output" / "consolidated_invoices.xlsx"
    csv_path = base_dir / "output" / "consolidated_invoices.csv"
    
    if not invoices_dir.exists():
        print(f"Error: Invoices directory not found at {invoices_dir}")
        return None, None, None
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        return None, None, None
    
    return invoices_dir, excel_path, csv_path

def extract_pdf_screenshot(pdf_path, dpi=150):
    """Extract a screenshot from PDF focusing on the main transaction area."""
    try:
        # Open PDF
        doc = fitz.open(pdf_path)
        
        # Get first page
        page = doc[0]
        
        # Render page to image at specified DPI
        mat = fitz.Matrix(dpi/72.0, dpi/72.0)
        pix = page.get_pixmap(matrix=mat)
        
        # Convert to PIL Image
        img_data = pix.pil_tobytes(format="PNG")
        img = Image.open(io.BytesIO(img_data))
        
        # Get dimensions
        width, height = img.size
        
        # Crop to focus on main content area (remove headers/footers)
        # Crop top 15% and bottom 10% to focus on transaction details
        crop_top = int(height * 0.15)
        crop_bottom = int(height * 0.90)
        crop_left = int(width * 0.05)
        crop_right = int(width * 0.95)
        
        # Crop the image
        img_cropped = img.crop((crop_left, crop_top, crop_right, crop_bottom))
        
        # Resize for Excel (max width 400px to fit nicely in cells)
        max_width = 400
        if img_cropped.width > max_width:
            ratio = max_width / img_cropped.width
            new_height = int(img_cropped.height * ratio)
            img_cropped = img_cropped.resize((max_width, new_height), Image.Resampling.LANCZOS)
        
        doc.close()
        
        return img_cropped
        
    except Exception as e:
        print(f"Error extracting screenshot from {pdf_path.name}: {str(e)}")
        return None

def add_screenshots_to_excel(excel_path, csv_path, invoices_dir):
    """Add screenshots to Excel file in a new sheet."""
    
    # Read transaction data from CSV
    df = pd.read_csv(csv_path)
    
    # Load existing Excel file
    wb = load_workbook(excel_path)
    
    # Remove sheet if it already exists
    if 'Invoice Screenshots' in wb.sheetnames:
        del wb['Invoice Screenshots']
    
    # Create new sheet
    ws = wb.create_sheet('Invoice Screenshots')
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['Invoice #', 'Date', 'Description', 'Amount (THB)', 'Source File', 'Screenshot']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Invoice #
    ws.column_dimensions['B'].width = 12  # Date
    ws.column_dimensions['C'].width = 40  # Description
    ws.column_dimensions['D'].width = 15  # Amount
    ws.column_dimensions['E'].width = 35  # Source File
    ws.column_dimensions['F'].width = 55  # Screenshot (400px ≈ 55 Excel units)
    
    # Process each transaction
    print("\nProcessing screenshots:")
    print("=" * 60)
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row (1-based, skip header)
        
        # Add text data
        ws.cell(row=row_num, column=1, value=idx + 1).border = border_style  # Invoice #
        ws.cell(row=row_num, column=2, value=row['Date']).border = border_style
        ws.cell(row=row_num, column=3, value=row['Description']).border = border_style
        
        # Format amount
        amount_cell = ws.cell(row=row_num, column=4, value=row['Amount (THB)'])
        amount_cell.border = border_style
        amount_cell.number_format = '#,##0.00'
        
        ws.cell(row=row_num, column=5, value=row['Source File']).border = border_style
        
        # Extract and add screenshot
        pdf_path = invoices_dir / row['Source File']
        
        if pdf_path.exists():
            print(f"[{idx+1}/{len(df)}] Processing {row['Source File']}...")
            
            # Extract screenshot
            img = extract_pdf_screenshot(pdf_path)
            
            if img:
                # Save to temporary BytesIO object
                img_buffer = io.BytesIO()
                img.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                
                # Create openpyxl Image object
                xl_img = XLImage(img_buffer)
                
                # Position image in cell F{row_num}
                cell_ref = f'F{row_num}'
                
                # Add image to worksheet
                ws.add_image(xl_img, cell_ref)
                
                # Set row height to accommodate image (approximately)
                # Image height in pixels / 0.75 = row height in Excel units
                ws.row_dimensions[row_num].height = img.height * 0.75
                
                print(f"  ✓ Screenshot added ({img.width}x{img.height}px)")
            else:
                ws.cell(row=row_num, column=6, value="Error loading image").border = border_style
                print(f"  ✗ Failed to extract screenshot")
        else:
            ws.cell(row=row_num, column=6, value="File not found").border = border_style
            print(f"  ✗ PDF file not found")
    
    # Add summary at the top of the sheet
    ws.insert_rows(1)
    summary_cell = ws.cell(row=1, column=1, value=f"Invoice Screenshots - {len(df)} transactions")
    summary_cell.font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Freeze panes (keep headers visible)
    ws.freeze_panes = 'A3'
    
    # Move sheet to position after Transactions
    if 'Transactions' in wb.sheetnames:
        sheets = wb.sheetnames
        transactions_idx = sheets.index('Transactions')
        screenshots_idx = sheets.index('Invoice Screenshots')
        
        # Reorder sheets
        wb.move_sheet('Invoice Screenshots', offset=transactions_idx - screenshots_idx + 1)
    
    # Save the workbook
    print("\nSaving Excel file...")
    wb.save(excel_path)
    
    return True

def main():
    """Main function to add PDF screenshots to Excel."""
    print("PDF to Excel Screenshots")
    print("=" * 60)
    
    # Setup paths
    invoices_dir, excel_path, csv_path = setup_paths()
    if not invoices_dir:
        sys.exit(1)
    
    # Count PDFs
    pdf_files = list(invoices_dir.glob("*.pdf"))
    print(f"Found {len(pdf_files)} PDF files to process")
    print(f"Excel file: {excel_path}")
    
    # Add screenshots to Excel
    if add_screenshots_to_excel(excel_path, csv_path, invoices_dir):
        print("\n" + "=" * 60)
        print("✓ Screenshots successfully added to Excel!")
        print(f"✓ Updated file: {excel_path}")
        print("✓ New sheet: 'Invoice Screenshots'")
        print("\nFeatures added:")
        print("  • Each invoice has its screenshot embedded")
        print("  • Screenshots are cropped to show main transaction area")
        print("  • Images are resized for optimal viewing in Excel")
        print("  • Complete visual reference for all transactions")
        print("=" * 60)
    else:
        print("\n✗ Failed to add screenshots to Excel")
        sys.exit(1)

if __name__ == "__main__":
    main()