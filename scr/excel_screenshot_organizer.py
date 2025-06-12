#!/usr/bin/env python3
"""
Excel Screenshot Organizer - One Image Per Sheet (Print-Optimized)

This script creates an Excel file with one screenshot per sheet,
sorted from oldest to newest based on the filename numbering.
Each sheet is optimized for single-page printing.
"""

import os
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from PIL import Image

def setup_directories():
    """Setup and validate directories."""
    base_dir = Path(__file__).parent.parent
    screenshots_dir = base_dir / "output" / "screenshot_zoomed"
    output_dir = base_dir / "output"
    
    if not screenshots_dir.exists():
        print(f"Error: Screenshots directory not found at {screenshots_dir}")
        return None, None
    
    return screenshots_dir, output_dir

def extract_number_from_filename(filename):
    """Extract the number from filename for sorting."""
    # Remove extension
    name = filename.stem
    
    # Handle the base file without number (treat as 0)
    if name == "American Express - Account Activity":
        return 0
    
    # Extract number from filename like "American Express - Account Activity-1"
    match = re.search(r'-(\d+(?:\.\d+)?)$', name)
    if match:
        return float(match.group(1))
    
    return 0

def get_sorted_screenshots(screenshots_dir):
    """Get all screenshot files sorted from oldest to newest."""
    png_files = list(screenshots_dir.glob("*.png"))
    
    if not png_files:
        print(f"No PNG files found in {screenshots_dir}")
        return []
    
    # Sort by extracted number (oldest to newest)
    sorted_files = sorted(png_files, key=extract_number_from_filename)
    
    print(f"Found {len(sorted_files)} screenshot files")
    return sorted_files

def setup_print_layout(ws):
    """Configure worksheet for single-page printing."""
    # Set page orientation to portrait
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    
    # Set paper size to A4 (8.27" x 11.69")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Set margins to very narrow for maximum space
    ws.page_margins = PageMargins(
        left=0.2,     # 0.2 inch
        right=0.2,    # 0.2 inch
        top=0.4,      # 0.4 inch (space for title)
        bottom=0.2,   # 0.2 inch
        header=0.2,
        footer=0.2
    )
    
    # Fit to one page
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # Set print quality
    ws.page_setup.horizontalDpi = 300
    ws.page_setup.verticalDpi = 300
    
    # Set view to Page Break Preview for better print visualization
    ws.sheet_view.view = "pageBreakPreview"

def calculate_print_optimized_size(img_width, img_height):
    """Calculate optimal image size for single-page printing."""
    # Available print area in portrait A4 size (in pixels at 96 DPI)
    # A4: 8.27" x 11.69" portrait, with minimal margins
    # Available area: ~7.8" x 11" = ~750px x 1056px at 96 DPI
    
    max_print_width = 750   # pixels (increased from 720)
    max_print_height = 1000 # pixels (increased from 950, leaving space for title)
    
    # Calculate scale factor to fit within print area
    width_scale = max_print_width / img_width
    height_scale = max_print_height / img_height
    scale_factor = min(width_scale, height_scale)
    
    # Apply scale factor
    display_width = int(img_width * scale_factor)
    display_height = int(img_height * scale_factor)
    
    return display_width, display_height

def create_excel_with_large_screenshots(screenshots, output_path):
    """Create Excel file with one print-optimized screenshot per sheet."""
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove the default sheet - we'll create our own
        wb.remove(wb.active)
        
        # Process each screenshot
        for i, screenshot_path in enumerate(screenshots):
            print(f"Processing sheet {i+1}/{len(screenshots)}: {screenshot_path.name}")
            
            # Create new worksheet for this screenshot
            sheet_name = f"Invoice_{i+1}"
            ws = wb.create_sheet(title=sheet_name)
            
            # Setup print layout first
            setup_print_layout(ws)
            
            # Add title with filename
            ws['A1'] = screenshot_path.stem
            ws['A1'].font = ws['A1'].font.copy(bold=True, size=12)
            ws.row_dimensions[1].height = 20
            
            # Set column widths optimized for A4 portrait printing
            ws.column_dimensions['A'].width = 90
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 15
            
            try:
                # Load and resize image for optimal printing
                with Image.open(screenshot_path) as img:
                    img_width, img_height = img.size
                    
                    # Calculate print-optimized size
                    display_width, display_height = calculate_print_optimized_size(img_width, img_height)
                    
                    print(f"  Image size: {img_width}x{img_height} -> Print-optimized: {display_width}x{display_height}")
                    
                    # Create Excel image object
                    excel_img = ExcelImage(screenshot_path)
                    
                    # Set image size for single-page printing
                    excel_img.width = display_width
                    excel_img.height = display_height
                    
                    # Position image starting from row 2 (right after title)
                    excel_img.anchor = 'A2'
                    ws.add_image(excel_img)
                    
                    # Set row heights to accommodate the image properly
                    # Calculate how many rows the image will span
                    rows_needed = max(1, int(display_height / 15))  # Approximate 15 pixels per row
                    for row in range(2, 2 + rows_needed):
                        ws.row_dimensions[row].height = min(100, max(15, display_height / rows_needed * 0.75))
                    
                    # Set fixed print area to columns A-B and rows 1-52
                    ws.print_area = 'A1:B52'
                
            except Exception as img_error:
                print(f"  Error processing image {screenshot_path.name}: {str(img_error)}")
                # Add error message to sheet
                ws['A2'] = f"Error loading image: {str(img_error)}"
                ws.print_area = 'A1:B52'
                continue
        
        # Save the workbook
        wb.save(output_path)
        print(f"\n✓ Excel file created successfully: {output_path}")
        return True
        
    except Exception as e:
        print(f"\n✗ Error creating Excel file: {str(e)}")
        return False

def main():
    """Main function to create Excel file with print-optimized screenshots."""
    print("Excel Screenshot Organizer - Print-Optimized Single Page")
    print("=" * 60)
    
    # Setup directories
    screenshots_dir, output_dir = setup_directories()
    if not screenshots_dir or not output_dir:
        sys.exit(1)
    
    # Get sorted screenshots
    screenshots = get_sorted_screenshots(screenshots_dir)
    if not screenshots:
        sys.exit(1)
    
    # Create output Excel file path
    excel_path = output_dir / "invoice_screenshots_organized.xlsx"
    
    # Create Excel file with print-optimized screenshots
    print(f"\nCreating Excel file with {len(screenshots)} print-optimized screenshots...")
    print("Format: One screenshot per sheet, optimized for single-page printing")
    
    if create_excel_with_large_screenshots(screenshots, excel_path):
        print(f"\n" + "=" * 60)
        print("✅ Excel file creation complete!")
        print(f"📁 File saved: {excel_path}")
        print(f"📊 Total screenshots: {len(screenshots)}")
        print(f"📋 Layout: One image per sheet ({len(screenshots)} sheets)")
        print("🖨️  Print-optimized: Each sheet fits on one page")
        print("📄 Page setup: A4 Portrait orientation with narrow margins")
        print("\nYou can now open the Excel file to view your organized screenshots!")
        print("Each sheet is configured for single-page printing in print preview.")
        print("Use the sheet tabs at the bottom to navigate between invoices.")
    else:
        print("\n❌ Failed to create Excel file")
        sys.exit(1)

if __name__ == "__main__":
    main() 